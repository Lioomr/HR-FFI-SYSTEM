"""Quick script to show Excel headers"""
from openpyxl import load_workbook

file_path = r"D:\FFI HR SYSTEM\HR FFI.xlsx"
workbook = load_workbook(file_path, read_only=True, data_only=True)
worksheet = workbook.active
header_row = list(next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), []))
workbook.close()

print(f"Total headers: {len(header_row)}\n")
for i, header in enumerate(header_row, 1):
    print(f"{i:2d}. '{header}'")
