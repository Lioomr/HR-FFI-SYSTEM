from datetime import date, timedelta
from io import BytesIO

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.core.exceptions import ValidationError as DjangoValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import IntegrityError, transaction
from openpyxl import Workbook, load_workbook
from rest_framework import status
from rest_framework.test import APITestCase

from announcements.models import Announcement
from assets.models import Asset
from attendance.models import AttendanceRecord, BioTimeDeviceEmployee, BioTimeEmployeeMap
from employees.management.commands.audit_employee_company_integrity import (
    build_employee_company_integrity_report,
)
from employees.models import EmployeeDocument, EmployeeImport, EmployeeProfile
from hr_reference.models import Department, Position
from in_app_notifications.models import Notification
from leaves.models import LeaveBalanceAdjustment, LeaveBalanceSnapshot, LeaveRequest, LeaveType
from leaves.utils import calculate_leave_balance
from organization.models import OrganizationNode, UserOrganizationAccess

User = get_user_model()


class TenantScopeRegressionTests(APITestCase):
    def setUp(self):
        self.client.defaults["HTTP_X_FORWARDED_PROTO"] = "https"
        self.hr_group, _ = Group.objects.get_or_create(name="HRManager")
        self.employee_group, _ = Group.objects.get_or_create(name="Employee")
        self.manager_group, _ = Group.objects.get_or_create(name="Manager")

        self.athroya = OrganizationNode.objects.create(
            code="ATHROYA_SCOPE_TEST",
            name="Athroya",
            node_type=OrganizationNode.NodeType.COMPANY,
            employee_id_prefix="ATH",
        )
        self.aseco = OrganizationNode.objects.create(
            code="ASECO_SCOPE_TEST",
            name="Aseco Pro",
            node_type=OrganizationNode.NodeType.COMPANY,
            employee_id_prefix="ASECO",
        )
        self.ffi = OrganizationNode.objects.create(
            code="FFI_SCOPE_TEST",
            name="FFI",
            node_type=OrganizationNode.NodeType.COMPANY,
            employee_id_prefix="FFI",
        )
        self.inactive_company = OrganizationNode.objects.create(
            code="INACTIVE_SCOPE_TEST",
            name="Inactive Company",
            node_type=OrganizationNode.NodeType.COMPANY,
            is_active=False,
        )

        self.athroya_hr = User.objects.create_user(email="athroya.hr@scope.test", password="password")
        self.athroya_hr.groups.add(self.hr_group)
        UserOrganizationAccess.objects.create(user=self.athroya_hr, organization=self.athroya)

        self.aseco_hr = User.objects.create_user(email="aseco.hr@scope.test", password="password")
        self.aseco_hr.groups.add(self.hr_group)
        UserOrganizationAccess.objects.create(user=self.aseco_hr, organization=self.aseco)

        self.athroya_user = User.objects.create_user(email="athroya.employee@scope.test", password="password")
        self.athroya_user.groups.add(self.employee_group)
        self.aseco_user = User.objects.create_user(email="aseco.employee@scope.test", password="password")
        self.ffi_user = User.objects.create_user(email="ffi.employee@scope.test", password="password")

        self.athroya_employee = self._employee(
            self.athroya_user,
            self.athroya,
            "ATH-SCOPE-001",
            "Athroya Employee",
        )
        self.aseco_employee = self._employee(
            self.aseco_user,
            self.aseco,
            "ASECO-SCOPE-001",
            "Aseco Employee",
        )
        self.ffi_employee = self._employee(
            self.ffi_user,
            self.ffi,
            "FFI-SCOPE-001",
            "FFI Employee",
        )
        self.null_company_employee = EmployeeProfile.objects.create(
            employee_id="NULL-SCOPE-001",
            full_name="Archived Unassigned Employee",
            hire_date=date(2020, 1, 1),
            is_archived=True,
        )

        self.athroya_leave_type = LeaveType.objects.create(
            company=self.athroya,
            name="Annual Leave",
            code="ANNUAL",
            annual_quota=21,
        )
        self.aseco_leave_type = LeaveType.objects.create(
            company=self.aseco,
            name="Annual Leave",
            code="ANNUAL",
            annual_quota=21,
        )
        LeaveRequest.objects.create(
            employee=self.athroya_user,
            employee_profile=self.athroya_employee,
            company=self.athroya,
            leave_type=self.athroya_leave_type,
            start_date=date(2026, 1, 10),
            end_date=date(2026, 1, 11),
            status=LeaveRequest.RequestStatus.APPROVED,
        )

    @staticmethod
    def _employee(user, company, employee_id, full_name, **kwargs):
        return EmployeeProfile.objects.create(
            user=user,
            company=company,
            employee_id=employee_id,
            full_name=full_name,
            hire_date=date(2020, 1, 1),
            employment_status=EmployeeProfile.EmploymentStatus.ACTIVE,
            **kwargs,
        )

    def _authenticate_athroya_hr(self):
        self.client.force_authenticate(user=self.athroya_hr)

    def _athroya_headers(self):
        return {"HTTP_X_ACTIVE_COMPANY_ID": str(self.athroya.id)}

    def test_company_employee_lists_are_bidirectionally_isolated_and_exclude_null(self):
        self._authenticate_athroya_hr()
        response = self.client.get("/api/employees/?page_size=50", **self._athroya_headers())

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        employee_ids = {item["employee_id"] for item in response.data["data"]["results"]}
        self.assertEqual(employee_ids, {"ATH-SCOPE-001"})
        self.assertEqual(response.data["data"]["count"], 1)

        self.client.force_authenticate(user=self.aseco_hr)
        response = self.client.get(
            "/api/employees/?page_size=50",
            HTTP_X_ACTIVE_COMPANY_ID=str(self.aseco.id),
        )
        employee_ids = {item["employee_id"] for item in response.data["data"]["results"]}
        self.assertEqual(employee_ids, {"ASECO-SCOPE-001"})
        self.assertNotIn("ATH-SCOPE-001", employee_ids)
        self.assertNotIn("FFI-SCOPE-001", employee_ids)

        archived = self.client.get(
            "/api/employees/?archive_state=archived&page_size=50",
            HTTP_X_ACTIVE_COMPANY_ID=str(self.aseco.id),
        )
        self.assertNotIn(
            "NULL-SCOPE-001",
            {item["employee_id"] for item in archived.data["data"]["results"]},
        )

    def test_unauthorized_or_invalid_company_selector_is_rejected(self):
        self._authenticate_athroya_hr()
        unauthorized = self.client.get(
            f"/api/employees/?company_id={self.aseco.id}",
            HTTP_X_ACTIVE_COMPANY_ID=str(self.aseco.id),
        )
        invalid = self.client.get("/api/employees/?company_id=not-an-id", **self._athroya_headers())
        blank = self.client.get("/api/employees/?company_id=", **self._athroya_headers())
        zero = self.client.get("/api/employees/?company_id=0", **self._athroya_headers())
        negative = self.client.get("/api/employees/?company_id=-1", **self._athroya_headers())
        duplicate = self.client.get(
            f"/api/employees/?company_id={self.athroya.id}&company_id={self.athroya.id}",
            **self._athroya_headers(),
        )
        conflicting = self.client.get(
            f"/api/employees/?company_id={self.athroya.id}",
            HTTP_X_ACTIVE_COMPANY_ID=str(self.aseco.id),
        )
        inactive = self.client.get(
            f"/api/employees/?company_id={self.inactive_company.id}",
            HTTP_X_ACTIVE_COMPANY_ID=str(self.inactive_company.id),
        )
        matching = self.client.get(
            f"/api/employees/?company_id={self.athroya.id}",
            **self._athroya_headers(),
        )
        unauthorized_header = self.client.get(
            "/api/employees/",
            HTTP_X_ACTIVE_COMPANY_ID=str(self.aseco.id),
        )

        self.assertEqual(unauthorized.status_code, status.HTTP_403_FORBIDDEN)
        self.assertEqual(invalid.status_code, status.HTTP_403_FORBIDDEN)
        self.assertEqual(blank.status_code, status.HTTP_403_FORBIDDEN)
        self.assertEqual(zero.status_code, status.HTTP_403_FORBIDDEN)
        self.assertEqual(negative.status_code, status.HTTP_403_FORBIDDEN)
        self.assertEqual(duplicate.status_code, status.HTTP_403_FORBIDDEN)
        self.assertEqual(conflicting.status_code, status.HTTP_403_FORBIDDEN)
        self.assertEqual(inactive.status_code, status.HTTP_403_FORBIDDEN)
        self.assertEqual(matching.status_code, status.HTTP_200_OK)
        self.assertEqual(unauthorized_header.status_code, status.HTTP_403_FORBIDDEN)

    def test_head_office_context_cannot_expand_an_ordinary_company_list(self):
        head_office = OrganizationNode.objects.create(
            code="HEAD_SCOPE_TEST",
            name="Head Office Scope Test",
            node_type=OrganizationNode.NodeType.HEAD_OFFICE,
        )
        UserOrganizationAccess.objects.create(user=self.athroya_hr, organization=head_office)
        self._authenticate_athroya_hr()

        response = self.client.get(
            "/api/employees/?page_size=50",
            HTTP_X_ACTIVE_COMPANY_ID=str(head_office.id),
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_employee_detail_export_and_pagination_share_the_same_scope(self):
        self._authenticate_athroya_hr()
        own_detail = self.client.get(
            f"/api/employees/{self.athroya_employee.id}/",
            **self._athroya_headers(),
        )
        foreign_detail = self.client.get(
            f"/api/employees/{self.aseco_employee.id}/",
            **self._athroya_headers(),
        )
        export = self.client.get("/api/employees/export/", **self._athroya_headers())

        self.assertEqual(own_detail.status_code, status.HTTP_200_OK)
        self.assertEqual(foreign_detail.status_code, status.HTTP_404_NOT_FOUND)
        self.assertEqual(export.status_code, status.HTTP_200_OK)
        workbook = load_workbook(BytesIO(export.content), read_only=True)
        exported_ids = {row[0] for row in list(workbook.active.iter_rows(values_only=True))[1:]}
        workbook.close()
        self.assertEqual(exported_ids, {"ATH-SCOPE-001"})

    def test_attendance_assets_documents_dashboard_and_notifications_are_scoped(self):
        AttendanceRecord.objects.create(
            employee_profile=self.athroya_employee,
            date=date.today(),
            status=AttendanceRecord.Status.PRESENT,
        )
        foreign_attendance = AttendanceRecord.objects.create(
            employee_profile=self.aseco_employee,
            date=date.today(),
            status=AttendanceRecord.Status.PRESENT,
        )
        foreign_leave = LeaveRequest.objects.create(
            employee=self.aseco_user,
            employee_profile=self.aseco_employee,
            company=self.aseco,
            leave_type=self.aseco_leave_type,
            start_date=date(2026, 6, 1),
            end_date=date(2026, 6, 2),
        )
        athroya_asset = Asset.objects.create(
            company=self.athroya,
            name_en="Athroya Asset",
            type=Asset.AssetType.OTHER,
            flexible_attributes={"kind": "test"},
        )
        Asset.objects.create(
            company=self.aseco,
            name_en="Aseco Asset",
            type=Asset.AssetType.OTHER,
            flexible_attributes={"kind": "test"},
        )
        document = EmployeeDocument.objects.create(
            employee_profile=self.athroya_employee,
            company=self.athroya,
            document_type=EmployeeDocument.DocumentType.OTHER,
            custom_name="Athroya Document",
            file=SimpleUploadedFile("athroya.pdf", b"%PDF-1.4\n%%EOF", content_type="application/pdf"),
        )
        EmployeeDocument.objects.create(
            employee_profile=self.aseco_employee,
            company=self.aseco,
            document_type=EmployeeDocument.DocumentType.OTHER,
            custom_name="Aseco Document",
            file=SimpleUploadedFile("aseco.pdf", b"%PDF-1.4\n%%EOF", content_type="application/pdf"),
        )
        athroya_notification = Notification.objects.create(
            recipient=self.athroya_hr,
            company=self.athroya,
            title="Athroya",
            event_key="scope.athroya",
        )
        Notification.objects.create(
            recipient=self.athroya_hr,
            company=self.aseco,
            title="Aseco",
            event_key="scope.aseco",
        )
        Notification.objects.create(
            recipient=self.athroya_hr,
            company=None,
            title="Unassigned",
            event_key="scope.null",
        )

        self._authenticate_athroya_hr()
        attendance = self.client.get("/api/attendance/?page_size=50", **self._athroya_headers())
        assets = self.client.get("/api/assets/?page_size=50", **self._athroya_headers())
        documents = self.client.get(
            f"/api/employees/{self.athroya_employee.id}/documents/",
            **self._athroya_headers(),
        )
        foreign_documents = self.client.get(
            f"/api/employees/{self.aseco_employee.id}/documents/",
            **self._athroya_headers(),
        )
        foreign_attendance_detail = self.client.get(
            f"/api/attendance/{foreign_attendance.id}/",
            **self._athroya_headers(),
        )
        foreign_leave_detail = self.client.get(
            f"/api/leaves/leave-requests/{foreign_leave.id}/",
            **self._athroya_headers(),
        )
        dashboard = self.client.get("/api/hr/summary/", **self._athroya_headers())
        notifications = self.client.get("/api/notifications/?page_size=50", **self._athroya_headers())

        self.assertEqual(attendance.data["data"]["count"], 1)
        self.assertEqual({item["id"] for item in assets.data["data"]["items"]}, {athroya_asset.id})
        self.assertEqual({item["id"] for item in documents.data["data"]}, {document.id})
        self.assertEqual(foreign_documents.status_code, status.HTTP_404_NOT_FOUND)
        self.assertIn(
            foreign_attendance_detail.status_code,
            {status.HTTP_403_FORBIDDEN, status.HTTP_404_NOT_FOUND},
        )
        self.assertIn(
            foreign_leave_detail.status_code,
            {status.HTTP_403_FORBIDDEN, status.HTTP_404_NOT_FOUND},
        )
        self.assertEqual(dashboard.data["data"]["total_employees"], 1)
        self.assertEqual(
            {item["id"] for item in notifications.data["data"]["items"]},
            {athroya_notification.id},
        )

    def test_reference_data_and_announcements_are_scoped(self):
        athroya_department = Department.objects.create(company=self.athroya, code="ATH-D", name="Athroya Dept")
        Department.objects.create(company=self.aseco, code="ASE-D", name="Aseco Dept")
        Department.objects.create(code="NULL-D", name="Unassigned Dept")
        athroya_position = Position.objects.create(company=self.athroya, code="ATH-P", name="Athroya Position")
        Position.objects.create(company=self.aseco, code="ASE-P", name="Aseco Position")
        athroya_announcement = Announcement.objects.create(
            company=self.athroya,
            title="Athroya Announcement",
            content="Athroya only",
            target_roles=[],
            created_by=self.athroya_hr,
        )
        Announcement.objects.create(
            company=self.aseco,
            title="Aseco Announcement",
            content="Aseco only",
            target_roles=[],
            created_by=self.athroya_hr,
        )
        Announcement.objects.create(
            company=None,
            title="Unassigned Announcement",
            content="No company",
            target_roles=[],
            created_by=self.athroya_hr,
        )

        self._authenticate_athroya_hr()
        departments = self.client.get("/api/hr/departments/", **self._athroya_headers())
        positions = self.client.get("/api/hr/positions/", **self._athroya_headers())
        announcements = self.client.get("/api/announcements/?page_size=50", **self._athroya_headers())

        self.assertEqual({item["id"] for item in departments.data["data"]}, {athroya_department.id})
        self.assertEqual({item["id"] for item in positions.data["data"]}, {athroya_position.id})
        self.assertEqual(
            {item["id"] for item in announcements.data["data"]["items"]},
            {athroya_announcement.id},
        )

    def test_leave_balances_are_company_scoped_and_unchanged_by_reads(self):
        before = calculate_leave_balance(self.athroya_user, 2026, profile=self.athroya_employee)
        request_count = LeaveRequest.objects.count()
        adjustment_count = LeaveBalanceAdjustment.objects.count()
        snapshot_count = LeaveBalanceSnapshot.objects.count()

        self._authenticate_athroya_hr()
        response = self.client.get(
            "/api/leaves/leave-balances/",
            {"employee_id": self.athroya_employee.id, "year": 2026},
            **self._athroya_headers(),
        )
        foreign = self.client.get(
            "/api/leaves/leave-balances/",
            {"employee_id": self.aseco_employee.id, "year": 2026},
            **self._athroya_headers(),
        )
        after = calculate_leave_balance(self.athroya_user, 2026, profile=self.athroya_employee)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(foreign.status_code, status.HTTP_404_NOT_FOUND)
        self.assertEqual(after, before)
        self.assertEqual(LeaveRequest.objects.count(), request_count)
        self.assertEqual(LeaveBalanceAdjustment.objects.count(), adjustment_count)
        self.assertEqual(LeaveBalanceSnapshot.objects.count(), snapshot_count)

    def test_cross_company_leave_type_cannot_be_assigned(self):
        self.client.force_authenticate(user=self.athroya_user)
        response = self.client.post(
            "/api/leaves/leave-requests/",
            {
                "leave_type": self.aseco_leave_type.id,
                "start_date": date.today() + timedelta(days=10),
                "end_date": date.today() + timedelta(days=11),
                "reason": "Cross-company attempt",
            },
            format="json",
            **self._athroya_headers(),
        )

        self.assertIn(response.status_code, {status.HTTP_400_BAD_REQUEST, status.HTTP_422_UNPROCESSABLE_ENTITY})
        self.assertFalse(
            LeaveRequest.objects.filter(
                employee_profile=self.athroya_employee,
                leave_type=self.aseco_leave_type,
            ).exists()
        )

    def test_manager_cannot_access_cross_company_employee(self):
        manager_user = User.objects.create_user(email="manager@scope.test", password="password")
        manager_user.groups.add(self.manager_group)
        manager_profile = self._employee(manager_user, self.athroya, "ATH-MANAGER", "Athroya Manager")
        with self.assertRaises(IntegrityError), transaction.atomic():
            EmployeeProfile.objects.filter(pk=self.aseco_employee.pk).update(
                manager_profile=manager_profile,
                manager=manager_user,
            )

        self.client.force_authenticate(user=manager_user)
        detail = self.client.get(
            f"/api/employees/{self.aseco_employee.id}/",
            **self._athroya_headers(),
        )
        team = self.client.get("/api/employees/manager/team/", **self._athroya_headers())

        self.assertIn(detail.status_code, {status.HTTP_403_FORBIDDEN, status.HTTP_404_NOT_FOUND})
        self.assertEqual(team.status_code, status.HTTP_403_FORBIDDEN)

    def test_company_integrity_validation_rejects_active_null_manager_and_balance_mismatches(self):
        active_null = EmployeeProfile(employee_id="ACTIVE-NULL", full_name="Invalid Active Null")
        with self.assertRaises(DjangoValidationError):
            active_null.full_clean()

        archived_legacy = EmployeeProfile(
            employee_id="ARCHIVED-NULL",
            full_name="Archived Legacy Employee",
            is_archived=True,
        )
        archived_legacy.full_clean()

        self.aseco_employee.manager_profile = self.athroya_employee
        with self.assertRaises(DjangoValidationError):
            self.aseco_employee.full_clean()

        cross_company_balance = LeaveBalanceSnapshot(
            employee_profile=self.athroya_employee,
            leave_type=self.aseco_leave_type,
            year=2026,
        )
        with self.assertRaises(DjangoValidationError):
            cross_company_balance.full_clean()

    def test_new_leave_request_resolves_company_from_employee_profile(self):
        request_record = LeaveRequest.objects.create(
            employee=self.athroya_user,
            leave_type=self.athroya_leave_type,
            start_date=date(2026, 4, 1),
            end_date=date(2026, 4, 2),
        )

        self.assertEqual(request_record.employee_profile_id, self.athroya_employee.id)
        self.assertEqual(request_record.company_id, self.athroya.id)

    def test_import_records_and_employees_receive_the_selected_company(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Emp Full Name", "Department", "Position Name", "Employee Number"])
        sheet.append(["Imported Athroya Employee", "Operations", "Engineer", "ATH-IMPORT-001"])
        payload = BytesIO()
        workbook.save(payload)
        upload = SimpleUploadedFile(
            "employees.xlsx",
            payload.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        self._authenticate_athroya_hr()
        response = self.client.post(
            "/api/employees/import/excel/",
            {"file": upload},
            format="multipart",
            **self._athroya_headers(),
        )

        self.assertEqual(response.status_code, status.HTTP_201_CREATED, response.data)
        imported = EmployeeProfile.objects.get(employee_number="ATH-IMPORT-001")
        self.assertEqual(imported.company_id, self.athroya.id)
        self.assertEqual(EmployeeImport.objects.get().company_id, self.athroya.id)

    def test_biotime_mapping_requires_an_employee_in_the_active_company(self):
        self._authenticate_athroya_hr()
        own = self.client.post(
            "/api/biotime-mappings/",
            {"employee_profile": self.athroya_employee.id, "biotime_emp_code": "ATH-BIO-1"},
            format="json",
            **self._athroya_headers(),
        )
        foreign = self.client.post(
            "/api/biotime-mappings/",
            {"employee_profile": self.aseco_employee.id, "biotime_emp_code": "ASE-BIO-1"},
            format="json",
            **self._athroya_headers(),
        )

        self.assertEqual(own.status_code, status.HTTP_201_CREATED, own.data)
        self.assertEqual(foreign.status_code, status.HTTP_422_UNPROCESSABLE_ENTITY)
        self.assertEqual(BioTimeEmployeeMap.objects.get().employee_profile.company_id, self.athroya.id)

        with self.assertRaises(DjangoValidationError):
            BioTimeEmployeeMap.objects.create(
                employee_profile=self.null_company_employee,
                biotime_emp_code="NULL-BIO-1",
            )

    def test_unscopable_biotime_device_roster_requires_explicit_all_company_head_office_context(self):
        BioTimeDeviceEmployee.objects.create(emp_code="GLOBAL-DEVICE-1", first_name="Global")
        self._authenticate_athroya_hr()

        tenant_response = self.client.get(
            "/api/biotime-mappings/unmapped/",
            **self._athroya_headers(),
        )
        self.assertEqual(tenant_response.status_code, status.HTTP_403_FORBIDDEN)

        system_admin = User.objects.create_user(email="biotime.admin@scope.test", password="password")
        system_admin.groups.add(Group.objects.get_or_create(name="SystemAdmin")[0])
        head_office = OrganizationNode.objects.create(
            code="HEAD_OFFICE_BIOTIME_SCOPE",
            name="BioTime Scope Head Office",
            node_type=OrganizationNode.NodeType.HEAD_OFFICE,
        )
        self.client.force_authenticate(system_admin)
        global_response = self.client.get(
            "/api/biotime-mappings/unmapped/",
            HTTP_X_ACTIVE_COMPANY_ID=str(head_office.id),
        )

        self.assertEqual(global_response.status_code, status.HTTP_200_OK)
        self.assertEqual(global_response.data["data"]["count"], 1)


class EmployeeCompanyAuditReportTests(APITestCase):
    def test_audit_report_is_read_only_and_preserves_attendance_linked_aryam(self):
        company = OrganizationNode.objects.create(
            code="AUDIT_SCOPE_TEST",
            name="Audit Company",
            node_type=OrganizationNode.NodeType.COMPANY,
        )
        aryam = EmployeeProfile.objects.create(
            company=company,
            employee_id="ARYAM-AUDIT-001",
            full_name="Aryam Ahmed Alghamdi",
        )
        AttendanceRecord.objects.create(
            employee_profile=aryam,
            date=date(2026, 3, 1),
            status=AttendanceRecord.Status.PRESENT,
        )
        null_profile = EmployeeProfile.objects.create(
            employee_id="NULL-AUDIT-001",
            full_name="Archived Legacy Unassigned",
            is_archived=True,
        )
        before = list(EmployeeProfile.objects.values_list("id", "company_id").order_by("id"))

        report = build_employee_company_integrity_report()

        after = list(EmployeeProfile.objects.values_list("id", "company_id").order_by("id"))
        self.assertEqual(after, before)
        self.assertEqual(report["protected_deleted_employee_ids_present"], [])
        self.assertEqual(report["null_company_profiles"][0]["employee_profile_id"], null_profile.id)
        self.assertEqual(report["aryam_ahmed_alghamdi_records"][0]["employee_profile_id"], aryam.id)
        self.assertEqual(report["aryam_ahmed_alghamdi_records"][0]["attendance_record_count"], 1)
