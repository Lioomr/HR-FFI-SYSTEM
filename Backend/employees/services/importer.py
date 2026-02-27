import csv
import hashlib
import io
import random
import re
import string
import uuid
import zipfile
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation

from django.core.files.base import ContentFile
from django.db import transaction

from employees.models import EmployeeImport, EmployeeProfile
from employees.storage import PrivateUploadStorage
from hr_reference.models import Department, Position, Sponsor, TaskGroup

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None


MAX_IMPORT_FILE_SIZE = 5 * 1024 * 1024
MAX_IMPORT_ROWS = 5000
ALLOWED_IMPORT_MIME_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}


@dataclass
class ImportExecutionResult:
    ok: bool
    status_code: int
    errors: list[str]
    inserted_rows: int = 0
    row_count: int = 0
    file_hash: str = ""
    result: str = "failed"


class EmployeeImporter:
    HEADER_ALIASES = {
        "full_name_en": ["emp full name", "full name", "full name en", "employee name"],
        "full_name_ar": ["full name ar", "employee name ar"],
        "employee_number": ["employee number", "employee number ", "emp no", "employee no", "employee id", "emp id"],
        "nationality_en": ["nationality", "nationality en"],
        "nationality_ar": ["nationality ar"],
        "is_saudi": ["is saudi", "saudi", "saudi employee"],
        "position_name_en": ["position name", "position", "job title", "job title en"],
        "position_name_ar": ["position name ar", "job title ar"],
        "passport_no": ["passport number", "passport no", "passport"],
        "passport_expiry": ["passport expiry", "passport expiry date"],
        "national_id": ["id", "national id", "iqama id"],
        "id_expiry": ["id expiry", "national id expiry", "iqama expiry"],
        "date_of_birth": ["date of birth", "dob"],
        "job_offer": ["job offer"],
        "hire_date": ["joining date", "join date", "hire date"],
        "contract_date": ["contract date"],
        "contract_expiry": ["contract expiry date", "contract expiry"],
        "task_group_name": ["task group name", "task group", "project"],
        "location_name": ["location", "project location", "site"],
        "health_card": ["health card"],
        "health_card_expiry": ["health card expiry"],
        "mobile": ["mobile number", "mobile", "phone"],
        "sponsor_code": ["sponsor code", "sponsor"],
        "basic_salary": ["basic salary"],
        "transportation_allowance": ["transportation allowance"],
        "accommodation_allowance": ["accommodation allowance"],
        "telephone_allowance": ["telephone allowance"],
        "petrol_allowance": ["petrol allowance"],
        "other_allowance": ["other allowance"],
        "total_salary": ["total salary"],
        "allowed_overtime": ["allowed overtime"],
        "department_name_en": ["department", "department name", "department en"],
        "department_name_ar": ["department ar", "department name ar"],
        "manager_ref": [
            "manager",
            "manager employee number",
            "manager id",
            "line manager",
            "direct manager",
            "direct manger",
        ],
    }

    DATE_FORMATS = (
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%m/%d/%Y",
        "%m-%d-%Y",
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
        "%d/%m/%Y %H:%M:%S",
        "%d-%m-%Y %H:%M:%S",
        "%m/%d/%Y %H:%M:%S",
        "%m-%d-%Y %H:%M:%S",
    )

    def __init__(self):
        self.storage = PrivateUploadStorage()

    @staticmethod
    def _generate_employee_id():
        return f"EMP-{''.join(random.choices(string.digits, k=6))}"

    def _generate_unique_employee_id(self, existing_ids, max_attempts=20):
        for _ in range(max_attempts):
            candidate = self._generate_employee_id()
            if candidate not in existing_ids:
                existing_ids.add(candidate)
                return candidate
        return None

    @staticmethod
    def _normalize_cell(value):
        if value is None:
            return ""
        if isinstance(value, str):
            return value.strip()
        return str(value).strip()

    def _normalize_header(self, value):
        text = self._normalize_cell(value).lower()
        # Normalize punctuation/odd separators so header matching is resilient.
        text = re.sub(r"[^a-z0-9]+", " ", text)
        return " ".join(text.split())

    def _has_xlsx_signature(self, uploaded_file):
        signature = uploaded_file.read(4)
        uploaded_file.seek(0)
        return signature in (b"PK\x03\x04", b"PK\x05\x06", b"PK\x07\x08")

    def _compute_file_hash(self, uploaded_file):
        hasher = hashlib.sha256()
        for chunk in uploaded_file.chunks():
            hasher.update(chunk)
        uploaded_file.seek(0)
        return hasher.hexdigest()

    def _write_errors_csv(self, errors_detail):
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["row", "column", "message"])
        for detail in errors_detail:
            writer.writerow([detail["row"], detail["column"], detail["message"]])
        return output.getvalue().encode("utf-8")

    def _parse_date_flexible(self, raw_value):
        raw = self._normalize_cell(raw_value)
        if not raw:
            return None, "", None
        if isinstance(raw_value, datetime):
            return raw_value.date(), raw, None
        # Common Excel text rendering includes a zeroed time part.
        if raw.endswith(" 00:00:00"):
            raw = raw[:10]
        for fmt in self.DATE_FORMATS:
            try:
                return datetime.strptime(raw, fmt).date(), raw, None
            except ValueError:
                continue
        return None, raw, f"Invalid date format: {raw}"

    def _parse_decimal(self, raw_value):
        value = self._normalize_cell(raw_value)
        if not value:
            return None, None
        cleaned = value.replace(",", "").replace("$", "").replace("SAR", "").strip()
        try:
            return Decimal(cleaned), None
        except InvalidOperation:
            return None, f"Invalid numeric value: {value}"

    def _to_bool(self, raw_value):
        value = self._normalize_cell(raw_value).lower()
        if value in {"1", "true", "yes", "y"}:
            return True, None
        if value in {"0", "false", "no", "n"}:
            return False, None
        if value == "":
            return None, None
        return None, f"Invalid boolean value: {raw_value}"

    def _build_reference_lookup(self, queryset):
        lookup = {}
        for obj in queryset:
            if obj.name:
                lookup[obj.name.strip().lower()] = obj
            if obj.code:
                lookup[obj.code.strip().lower()] = obj
        return lookup

    def _header_index_map(self, header_row):
        normalized_headers = [self._normalize_header(value) for value in header_row]
        header_map = {}
        normalized_aliases = {
            canonical: {self._normalize_header(alias) for alias in aliases}
            for canonical, aliases in self.HEADER_ALIASES.items()
        }
        for canonical, aliases in self.HEADER_ALIASES.items():
            alias_set = normalized_aliases[canonical]
            for idx, header_value in enumerate(normalized_headers):
                if header_value in alias_set:
                    header_map[canonical] = idx
                    break
        return header_map

    def _store_import_record(self, uploader, upload, file_hash, row_count):
        upload.seek(0)
        stored_name = self.storage.save(f"employee_imports/{uuid.uuid4().hex}.xlsx", upload)
        return EmployeeImport.objects.create(
            uploader=uploader,
            original_filename=upload.name,
            stored_file=stored_name,
            status=EmployeeImport.Status.FAILED,
            row_count=row_count,
            file_hash=file_hash,
        )

    def _save_import_failure(self, record, errors, errors_detail):
        errors_content = self._write_errors_csv(errors_detail)
        record.errors_file.save(f"employee_imports/errors/{uuid.uuid4().hex}.csv", ContentFile(errors_content))
        record.error_summary = errors[:10]
        record.save(update_fields=["error_summary", "updated_at"])

    def _extract_by_header(self, values, header_map, key):
        idx = header_map.get(key)
        if idx is None or idx >= len(values):
            return ""
        return self._normalize_cell(values[idx])

    def _resolve_or_create_department(self, lookup, department_name):
        if not department_name:
            return None
        key = department_name.lower()
        if key in lookup:
            return lookup[key]
        department = Department.objects.filter(name=department_name).first()
        if not department:
            code = department_name[:10].upper().replace(" ", "_")
            department = Department.objects.filter(code=code).first()
            if not department:
                department = Department.objects.create(
                    name=department_name, code=code, description="Auto-created from import"
                )
        lookup[key] = department
        if department.code:
            lookup[department.code.lower()] = department
        return department

    def _resolve_or_create_position(self, lookup, position_name):
        if not position_name:
            return None
        key = position_name.lower()
        if key in lookup:
            return lookup[key]
        position = Position.objects.filter(name=position_name).first()
        if not position:
            code = position_name[:10].upper().replace(" ", "_")
            position = Position.objects.filter(code=code).first()
            if not position:
                position = Position.objects.create(
                    name=position_name, code=code, description="Auto-created from import"
                )
        lookup[key] = position
        if position.code:
            lookup[position.code.lower()] = position
        return position

    def _resolve_or_create_sponsor(self, lookup, sponsor_code):
        if not sponsor_code:
            return None
        key = sponsor_code.lower()
        if key in lookup:
            return lookup[key]
        sponsor = Sponsor.objects.filter(code__iexact=sponsor_code).first()
        if not sponsor:
            sponsor = Sponsor.objects.filter(name__iexact=sponsor_code).first()
        if not sponsor:
            sponsor = Sponsor.objects.create(
                name=sponsor_code,
                code=sponsor_code.upper().replace(" ", "_")[:20],
            )
        lookup[key] = sponsor
        if sponsor.code:
            lookup[sponsor.code.lower()] = sponsor
        if sponsor.name:
            lookup[sponsor.name.lower()] = sponsor
        return sponsor

    def _resolve_or_create_task_group(self, lookup, project_name, location_name):
        project_name = (project_name or "").strip()
        location_name = (location_name or "").strip()
        if not project_name and not location_name:
            return None

        # Prefer a combined label so both Project and Location are visible in Task Groups UI.
        if project_name and location_name:
            display_name = f"{project_name} - {location_name}"
        else:
            display_name = project_name or location_name

        key = display_name.lower()
        if key in lookup:
            return lookup[key]

        task_group = TaskGroup.objects.filter(name__iexact=display_name).first()
        if not task_group and project_name:
            # Fallback for old rows where only project was stored as task group name.
            task_group = TaskGroup.objects.filter(name__iexact=project_name).first()

        if not task_group:
            raw_code = f"{project_name}_{location_name}" if location_name else (project_name or location_name)
            code = raw_code.upper().replace(" ", "_")[:20] or "TASK_GROUP"
            task_group = TaskGroup.objects.create(
                name=display_name,
                code=code,
                description=f"Auto-created from import (Project: {project_name or '-'}, Location: {location_name or '-'})",
            )

        lookup[key] = task_group
        if task_group.name:
            lookup[task_group.name.lower()] = task_group
        if task_group.code:
            lookup[task_group.code.lower()] = task_group
        return task_group

    def execute(self, upload, uploader):
        file_hash = ""
        if upload is None:
            return ImportExecutionResult(ok=False, status_code=400, errors=["File is required."], result="failed")
        if upload.size > MAX_IMPORT_FILE_SIZE:
            return ImportExecutionResult(ok=False, status_code=413, errors=["File exceeds 5MB limit."], result="failed")
        if not upload.name.lower().endswith(".xlsx"):
            return ImportExecutionResult(ok=False, status_code=415, errors=["Unsupported file type."], result="failed")
        content_type = upload.content_type or ""
        if content_type not in ALLOWED_IMPORT_MIME_TYPES:
            return ImportExecutionResult(ok=False, status_code=415, errors=["Unsupported file type."], result="failed")
        if not self._has_xlsx_signature(upload):
            return ImportExecutionResult(ok=False, status_code=415, errors=["Unsupported file type."], result="failed")

        try:
            with zipfile.ZipFile(upload) as zf:
                if "xl/workbook.xml" not in zf.namelist():
                    return ImportExecutionResult(
                        ok=False, status_code=415, errors=["Unsupported file type."], result="failed"
                    )
            upload.seek(0)
        except zipfile.BadZipFile:
            return ImportExecutionResult(ok=False, status_code=415, errors=["Unsupported file type."], result="failed")

        if load_workbook is None:
            return ImportExecutionResult(
                ok=False, status_code=500, errors=["Excel parser unavailable."], result="failed"
            )

        file_hash = self._compute_file_hash(upload)
        try:
            workbook = load_workbook(upload, read_only=True, data_only=True)
        except Exception:
            return ImportExecutionResult(
                ok=False, status_code=415, errors=["Unsupported file type."], file_hash=file_hash, result="failed"
            )

        worksheet = workbook.active
        header_row = list(next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), []))
        while header_row and self._normalize_cell(header_row[-1]) == "":
            header_row.pop()

        header_map = self._header_index_map(header_row)
        required = ["full_name_en", "department_name_en", "position_name_en"]
        missing = [key for key in required if key not in header_map]
        if missing:
            workbook.close()
            record = self._store_import_record(uploader, upload, file_hash, 0)
            errors = [f"row 1: Missing required headers: {', '.join(missing)}."]
            detail = [{"row": 1, "column": "Header", "message": errors[0]}]
            self._save_import_failure(record, errors, detail)
            return ImportExecutionResult(
                ok=False, status_code=422, errors=errors, row_count=0, file_hash=file_hash, result="failed"
            )

        rows = []
        for row_index, row_values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            row_list = list(row_values or [])
            if all(not self._normalize_cell(value) for value in row_list):
                continue
            rows.append({"row_index": row_index, "values": row_list})

        workbook.close()
        row_count = len(rows)
        if row_count == 0:
            record = self._store_import_record(uploader, upload, file_hash, 0)
            errors = ["row 0: No data rows found."]
            detail = [{"row": 0, "column": "Sheet", "message": "No data rows found."}]
            self._save_import_failure(record, errors, detail)
            return ImportExecutionResult(
                ok=False, status_code=422, errors=errors, row_count=0, file_hash=file_hash, result="failed"
            )
        if row_count > MAX_IMPORT_ROWS:
            record = self._store_import_record(uploader, upload, file_hash, row_count)
            errors = [f"row 0: Row limit exceeded (max {MAX_IMPORT_ROWS})."]
            detail = [{"row": 0, "column": "Sheet", "message": errors[0]}]
            self._save_import_failure(record, errors, detail)
            return ImportExecutionResult(
                ok=False, status_code=422, errors=errors, row_count=row_count, file_hash=file_hash, result="failed"
            )

        department_lookup = self._build_reference_lookup(Department.objects.all())
        position_lookup = self._build_reference_lookup(Position.objects.all())
        task_group_lookup = self._build_reference_lookup(TaskGroup.objects.all())
        sponsor_lookup = self._build_reference_lookup(Sponsor.objects.all())

        errors = []
        errors_detail = []
        prepared_rows = []

        for row in rows:
            row_index = row["row_index"]
            values = row["values"]
            row_has_error = False

            full_name_en = self._extract_by_header(values, header_map, "full_name_en")
            full_name_ar = self._extract_by_header(values, header_map, "full_name_ar")
            employee_number = self._extract_by_header(values, header_map, "employee_number")
            nationality_en = self._extract_by_header(values, header_map, "nationality_en")
            nationality_ar = self._extract_by_header(values, header_map, "nationality_ar")
            position_name_en = self._extract_by_header(values, header_map, "position_name_en")
            position_name_ar = self._extract_by_header(values, header_map, "position_name_ar")
            passport_no = self._extract_by_header(values, header_map, "passport_no")
            national_id = self._extract_by_header(values, header_map, "national_id")
            job_offer = self._extract_by_header(values, header_map, "job_offer")
            task_group_name = self._extract_by_header(values, header_map, "task_group_name")
            location_name = self._extract_by_header(values, header_map, "location_name")
            health_card = self._extract_by_header(values, header_map, "health_card")
            mobile = self._extract_by_header(values, header_map, "mobile")
            sponsor_code = self._extract_by_header(values, header_map, "sponsor_code")
            department_name_en = self._extract_by_header(values, header_map, "department_name_en")
            department_name_ar = self._extract_by_header(values, header_map, "department_name_ar")
            manager_ref = self._extract_by_header(values, header_map, "manager_ref")

            if not full_name_en:
                errors.append(f"row {row_index}: full_name_en is required.")
                errors_detail.append({"row": row_index, "column": "full_name_en", "message": "Required field."})
                row_has_error = True
            is_saudi, bool_err = self._to_bool(self._extract_by_header(values, header_map, "is_saudi"))
            if bool_err:
                errors.append(f"row {row_index}: {bool_err}")
                errors_detail.append({"row": row_index, "column": "is_saudi", "message": bool_err})
                row_has_error = True
            if is_saudi is None:
                is_saudi = "saudi" in nationality_en.lower()

            passport_expiry, passport_expiry_raw, err = self._parse_date_flexible(
                self._extract_by_header(values, header_map, "passport_expiry")
            )
            if err:
                errors.append(f"row {row_index}: {err}")
                errors_detail.append({"row": row_index, "column": "passport_expiry", "message": err})
                row_has_error = True

            id_expiry, id_expiry_raw, err = self._parse_date_flexible(
                self._extract_by_header(values, header_map, "id_expiry")
            )
            if err:
                errors.append(f"row {row_index}: {err}")
                errors_detail.append({"row": row_index, "column": "id_expiry", "message": err})
                row_has_error = True

            date_of_birth, date_of_birth_raw, err = self._parse_date_flexible(
                self._extract_by_header(values, header_map, "date_of_birth")
            )
            if err:
                errors.append(f"row {row_index}: {err}")
                errors_detail.append({"row": row_index, "column": "date_of_birth", "message": err})
                row_has_error = True

            hire_date, hire_date_raw, err = self._parse_date_flexible(
                self._extract_by_header(values, header_map, "hire_date")
            )
            if err:
                errors.append(f"row {row_index}: {err}")
                errors_detail.append({"row": row_index, "column": "hire_date", "message": err})
                row_has_error = True

            contract_date, contract_date_raw, err = self._parse_date_flexible(
                self._extract_by_header(values, header_map, "contract_date")
            )
            if err:
                errors.append(f"row {row_index}: {err}")
                errors_detail.append({"row": row_index, "column": "contract_date", "message": err})
                row_has_error = True

            contract_expiry, contract_expiry_raw, err = self._parse_date_flexible(
                self._extract_by_header(values, header_map, "contract_expiry")
            )
            if err:
                errors.append(f"row {row_index}: {err}")
                errors_detail.append({"row": row_index, "column": "contract_expiry", "message": err})
                row_has_error = True

            health_card_expiry, health_card_expiry_raw, err = self._parse_date_flexible(
                self._extract_by_header(values, header_map, "health_card_expiry")
            )
            if err:
                errors.append(f"row {row_index}: {err}")
                errors_detail.append({"row": row_index, "column": "health_card_expiry", "message": err})
                row_has_error = True

            basic_salary, err = self._parse_decimal(self._extract_by_header(values, header_map, "basic_salary"))
            if err:
                errors.append(f"row {row_index}: {err}")
                errors_detail.append({"row": row_index, "column": "basic_salary", "message": err})
                row_has_error = True

            transportation_allowance, err = self._parse_decimal(
                self._extract_by_header(values, header_map, "transportation_allowance")
            )
            if err:
                errors.append(f"row {row_index}: {err}")
                errors_detail.append({"row": row_index, "column": "transportation_allowance", "message": err})
                row_has_error = True

            accommodation_allowance, err = self._parse_decimal(
                self._extract_by_header(values, header_map, "accommodation_allowance")
            )
            if err:
                errors.append(f"row {row_index}: {err}")
                errors_detail.append({"row": row_index, "column": "accommodation_allowance", "message": err})
                row_has_error = True

            telephone_allowance, err = self._parse_decimal(
                self._extract_by_header(values, header_map, "telephone_allowance")
            )
            if err:
                errors.append(f"row {row_index}: {err}")
                errors_detail.append({"row": row_index, "column": "telephone_allowance", "message": err})
                row_has_error = True

            petrol_allowance, err = self._parse_decimal(self._extract_by_header(values, header_map, "petrol_allowance"))
            if err:
                errors.append(f"row {row_index}: {err}")
                errors_detail.append({"row": row_index, "column": "petrol_allowance", "message": err})
                row_has_error = True

            other_allowance, err = self._parse_decimal(self._extract_by_header(values, header_map, "other_allowance"))
            if err:
                errors.append(f"row {row_index}: {err}")
                errors_detail.append({"row": row_index, "column": "other_allowance", "message": err})
                row_has_error = True

            total_salary, err = self._parse_decimal(self._extract_by_header(values, header_map, "total_salary"))
            if err:
                errors.append(f"row {row_index}: {err}")
                errors_detail.append({"row": row_index, "column": "total_salary", "message": err})
                row_has_error = True

            allowed_overtime = None
            allowed_overtime_raw = self._extract_by_header(values, header_map, "allowed_overtime")
            if allowed_overtime_raw:
                normalized_allowed = allowed_overtime_raw.strip().lower()
                if normalized_allowed in {"false", "no", "n"}:
                    allowed_overtime = 0
                elif normalized_allowed in {"true", "yes", "y"}:
                    allowed_overtime = 1
                else:
                    try:
                        allowed_overtime = int(allowed_overtime_raw)
                    except ValueError:
                        msg = "Invalid numeric value."
                        errors.append(f"row {row_index}: {msg}")
                        errors_detail.append({"row": row_index, "column": "allowed_overtime", "message": msg})
                        row_has_error = True

            # Keep import permissive: do not fail row when identity docs are missing.
            # For Saudi profiles, passport fields are ignored.
            if is_saudi:
                passport_no = None
                passport_expiry = None
                passport_expiry_raw = ""

            department_ref = self._resolve_or_create_department(department_lookup, department_name_en)
            position_ref = self._resolve_or_create_position(position_lookup, position_name_en)
            task_group_ref = self._resolve_or_create_task_group(task_group_lookup, task_group_name, location_name)
            sponsor_ref = self._resolve_or_create_sponsor(sponsor_lookup, sponsor_code)

            if not row_has_error:
                prepared_rows.append(
                    {
                        "full_name": full_name_en,
                        "full_name_en": full_name_en,
                        "full_name_ar": full_name_ar,
                        "employee_number": employee_number,
                        "nationality": nationality_en,
                        "nationality_en": nationality_en,
                        "nationality_ar": nationality_ar,
                        "is_saudi": is_saudi,
                        "passport_no": passport_no,
                        "passport_expiry": passport_expiry,
                        "passport_expiry_raw": passport_expiry_raw,
                        "national_id": national_id,
                        "id_expiry": id_expiry,
                        "id_expiry_raw": id_expiry_raw,
                        "date_of_birth": date_of_birth,
                        "date_of_birth_raw": date_of_birth_raw,
                        "job_offer": job_offer,
                        "hire_date": hire_date,
                        "hire_date_raw": hire_date_raw,
                        "contract_date": contract_date,
                        "contract_date_raw": contract_date_raw,
                        "contract_expiry": contract_expiry,
                        "contract_expiry_raw": contract_expiry_raw,
                        "task_group_ref": task_group_ref,
                        "health_card": health_card,
                        "health_card_expiry": health_card_expiry,
                        "health_card_expiry_raw": health_card_expiry_raw,
                        "mobile": mobile,
                        "sponsor_ref": sponsor_ref,
                        "basic_salary": basic_salary,
                        "transportation_allowance": transportation_allowance,
                        "accommodation_allowance": accommodation_allowance,
                        "telephone_allowance": telephone_allowance,
                        "petrol_allowance": petrol_allowance,
                        "other_allowance": other_allowance,
                        "total_salary": total_salary,
                        "allowed_overtime": allowed_overtime,
                        "department_ref": department_ref,
                        "position_ref": position_ref,
                        "department": department_name_en or "",
                        "department_name_en": department_name_en or "",
                        "department_name_ar": department_name_ar or "",
                        "job_title": position_name_en or "",
                        "job_title_en": position_name_en or "",
                        "job_title_ar": position_name_ar or "",
                        "manager_ref": manager_ref,
                        "data_source": EmployeeProfile.DataSource.IMPORT_EXCEL,
                    }
                )

        import_record = self._store_import_record(uploader, upload, file_hash, row_count)
        if errors:
            self._save_import_failure(import_record, errors, errors_detail)
            return ImportExecutionResult(
                ok=False,
                status_code=422,
                errors=errors,
                inserted_rows=0,
                row_count=row_count,
                file_hash=file_hash,
                result="failed",
            )

        employee_numbers = [row["employee_number"] for row in prepared_rows if row["employee_number"]]
        national_ids = [row["national_id"] for row in prepared_rows if row["national_id"]]
        passport_numbers = [row["passport_no"] for row in prepared_rows if row["passport_no"]]
        existing_profiles_by_number = {
            ep.employee_number: ep
            for ep in EmployeeProfile.objects.filter(employee_number__in=employee_numbers).select_related(
                "manager_profile"
            )
        }
        existing_profiles_by_national_id = {
            ep.national_id: ep
            for ep in EmployeeProfile.objects.filter(national_id__in=national_ids).select_related("manager_profile")
            if ep.national_id
        }
        existing_profiles_by_passport = {
            ep.passport_no: ep
            for ep in EmployeeProfile.objects.filter(passport_no__in=passport_numbers).select_related("manager_profile")
            if ep.passport_no
        }
        all_profiles_by_emp_no = {
            ep.employee_number: ep
            for ep in EmployeeProfile.objects.exclude(employee_number="").only("id", "employee_number")
        }

        created_or_updated = {}
        manager_links = []
        existing_ids = set(EmployeeProfile.objects.values_list("employee_id", flat=True))
        try:
            with transaction.atomic():
                for row in prepared_rows:
                    manager_ref = row.pop("manager_ref", "")
                    employee_number = row["employee_number"]
                    profile = None
                    if employee_number:
                        profile = existing_profiles_by_number.get(employee_number)
                    if profile is None and row["national_id"]:
                        profile = existing_profiles_by_national_id.get(row["national_id"])
                    if profile is None and row["passport_no"]:
                        profile = existing_profiles_by_passport.get(row["passport_no"])
                    if profile is None:
                        employee_id = self._generate_unique_employee_id(existing_ids)
                        if not employee_id:
                            raise ValueError("Failed to generate unique employee_id.")
                        profile = EmployeeProfile.objects.create(employee_id=employee_id, **row)
                        if employee_number:
                            existing_profiles_by_number[employee_number] = profile
                        if profile.national_id:
                            existing_profiles_by_national_id[profile.national_id] = profile
                        if profile.passport_no:
                            existing_profiles_by_passport[profile.passport_no] = profile
                    else:
                        if not row["employee_number"]:
                            row.pop("employee_number")
                        for field, value in row.items():
                            setattr(profile, field, value)
                        profile.save()
                    if employee_number:
                        created_or_updated[employee_number] = profile
                        all_profiles_by_emp_no[employee_number] = profile
                    manager_links.append((profile, manager_ref))

                for profile, manager_ref in manager_links:
                    if not manager_ref:
                        profile.manager_profile = None
                        profile.manager = None
                        profile.save(update_fields=["manager_profile", "manager", "updated_at"])
                        continue
                    manager_profile = created_or_updated.get(manager_ref) or all_profiles_by_emp_no.get(manager_ref)
                    if manager_profile is None:
                        manager_profile = (
                            EmployeeProfile.objects.filter(full_name_en__iexact=manager_ref).only("id", "user").first()
                            or EmployeeProfile.objects.filter(full_name__iexact=manager_ref).only("id", "user").first()
                        )
                    if manager_profile and manager_profile.id != profile.id:
                        profile.manager_profile = manager_profile
                        profile.manager = manager_profile.user
                        profile.save(update_fields=["manager_profile", "manager", "updated_at"])
        except Exception:
            errors = ["Import failed due to a database constraint."]
            detail = [{"row": 0, "column": "Database", "message": "Constraint violation."}]
            self._save_import_failure(import_record, errors, detail)
            return ImportExecutionResult(
                ok=False,
                status_code=422,
                errors=errors,
                inserted_rows=0,
                row_count=row_count,
                file_hash=file_hash,
                result="failed",
            )

        import_record.status = EmployeeImport.Status.SUCCESS
        import_record.inserted_rows = row_count
        import_record.error_summary = []
        import_record.save(update_fields=["status", "inserted_rows", "error_summary", "updated_at"])
        return ImportExecutionResult(
            ok=True,
            status_code=201,
            errors=[],
            inserted_rows=row_count,
            row_count=row_count,
            file_hash=file_hash,
            result="success",
        )
